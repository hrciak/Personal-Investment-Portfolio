import openpyxl
from parsers.base import create_transaction

def strip_ticker(ticker: str) -> str:
    """Strips exchange suffix from XTB tickers (e.g. FB.US -> FB)"""
    if not ticker:
        return ""
    return ticker.split('.')[0]

def find_header_row(sheet, required_headers: list[str]) -> int:
    """Finds the 1-based index of the header row in a sheet."""
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if not row:
            continue
        row_strs = [str(cell).strip() for cell in row if cell is not None]
        if all(any(req in cell_str for cell_str in row_strs) for req in required_headers):
            return row_idx
    return -1

def extract_rows_as_dicts(sheet, header_row_idx: int) -> list[dict]:
    rows = list(sheet.iter_rows(values_only=True))
    header = [str(c).strip() if c is not None else "" for c in rows[header_row_idx - 1]]
    
    data = []
    for row in rows[header_row_idx:]:
        if not any(row):  # skip empty
            continue
        row_dict = {}
        for col_idx, col_name in enumerate(header):
            if col_idx < len(row):
                row_dict[col_name] = row[col_idx]
        data.append(row_dict)
    return data

def parse_xtb_xlsx(filepath: str) -> list[dict]:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    transactions = []
    
    # 1. CLOSED POSITION HISTORY
    if "CLOSED POSITION HISTORY" in wb.sheetnames:
        sheet = wb["CLOSED POSITION HISTORY"]
        hdr_idx = find_header_row(sheet, ["Symbol", "Open time"])
        if hdr_idx != -1:
            rows = extract_rows_as_dicts(sheet, hdr_idx)
            for r in rows:
                if not r.get("Symbol") or not r.get("Open time") or not r.get("Close time"):
                    continue
                
                raw_symbol = str(r.get("Symbol", ""))
                ticker = strip_ticker(raw_symbol)
                vol = float(r.get("Volume", 0.0))
                open_p = float(r.get("Open price", 0.0))
                close_p = float(r.get("Close price", 0.0))
                comm = float(r.get("Commission", 0.0))
                pos_id = str(r.get("Position", ""))
                gross_pl = float(r.get("Gross P/L", 0.0) or r.get("Gross P/L ", 0.0))

                # BUY leg
                transactions.append(create_transaction(
                    date=r["Open time"],
                    ticker=ticker,
                    tx_type="BUY",
                    qty=vol,
                    price=open_p,
                    fee=abs(comm) / 2,
                    realized_pnl=None,
                    position_id=pos_id,
                    source="XTB",
                    raw_symbol=raw_symbol
                ))

                # SELL leg
                transactions.append(create_transaction(
                    date=r["Close time"],
                    ticker=ticker,
                    tx_type="SELL",
                    qty=vol,
                    price=close_p,
                    fee=abs(comm) / 2,
                    realized_pnl=gross_pl,
                    position_id=pos_id,
                    source="XTB",
                    raw_symbol=raw_symbol
                ))

    # 2. OPEN POSITION
    open_pos_sheet_name = next((sn for sn in wb.sheetnames if "OPEN POSITION" in sn), None)
    if open_pos_sheet_name:
        sheet = wb[open_pos_sheet_name]
        hdr_idx = find_header_row(sheet, ["Symbol", "Open time", "Market price"])
        if hdr_idx != -1:
            rows = extract_rows_as_dicts(sheet, hdr_idx)
            for r in rows:
                if not r.get("Symbol") or not r.get("Open time"):
                    continue
                
                raw_symbol = str(r.get("Symbol", ""))
                ticker = strip_ticker(raw_symbol)
                vol = float(r.get("Volume", 0.0))
                open_p = float(r.get("Open price", 0.0))
                mkt_p = float(r.get("Market price", 0.0))
                pos_id = str(r.get("Position", ""))

                transactions.append(create_transaction(
                    date=r["Open time"],
                    ticker=ticker,
                    tx_type="BUY",
                    qty=vol,
                    price=open_p,
                    fee=0.0,
                    market_price_at_export=mkt_p,
                    position_id=pos_id,
                    source="XTB",
                    raw_symbol=raw_symbol
                ))

    # 3. CASH OPERATION HISTORY
    if "CASH OPERATION HISTORY" in wb.sheetnames:
        sheet = wb["CASH OPERATION HISTORY"]
        hdr_idx = find_header_row(sheet, ["Type", "Amount", "Time"])
        if hdr_idx != -1:
            rows = extract_rows_as_dicts(sheet, hdr_idx)
            for r in rows:
                op_type = str(r.get("Type", "")).lower()
                if not op_type or not r.get("Time"):
                    continue
                
                amt = float(r.get("Amount", 0.0))
                sym = strip_ticker(r.get("Symbol", ""))
                dt = r["Time"]

                if "deposit" in op_type:
                    transactions.append(create_transaction(date=dt, ticker="CASH", tx_type="DEPOSIT", qty=abs(amt), price=1.0, source="XTB"))
                elif "withdrawal" in op_type:
                    transactions.append(create_transaction(date=dt, ticker="CASH", tx_type="WITHDRAWAL", qty=abs(amt), price=1.0, source="XTB"))
                elif "divident" in op_type or "dividend" in op_type:
                    transactions.append(create_transaction(date=dt, ticker=sym, tx_type="DIVIDEND", qty=abs(amt), price=1.0, source="XTB"))
                elif "withholding tax" in op_type:
                    transactions.append(create_transaction(date=dt, ticker=sym, tx_type="WITHHOLDING_TAX", qty=abs(amt), price=1.0, source="XTB"))

    return transactions
