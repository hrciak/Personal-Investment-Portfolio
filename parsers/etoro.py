"""
eToro Account Statement Parser
==============================

eToro exports an XLSX "Account Statement" workbook. The sheet and column names
are localized to the account's language. This parser supports both English and
Czech exports (the two formats seen in practice) and is structured so more
locales can be added by extending the keyword tables below.

eToro accounts are denominated in USD, but the statement already provides EUR
figures for the values that matter (realized profit per closed position, and
the original EUR amount inside deposit descriptions). We use those EUR figures
directly and only fall back to an FX conversion for USD-only amounts. The FX
rate is derived from the statement itself (EUR shown in a deposit vs. its USD
amount) so it reflects the real period rate, with a live rate as a last resort.

Sheets consumed:
  - Closed Positions  / "Zavřené pozice"      -> BUY+SELL legs, EUR realized P/L
  - Open Positions    / "Otevřené pozice"     -> BUY legs for current holdings
  - Account Activity  / "Aktivita na účtu"    -> deposits, withdrawals, dividends
  - Dividends         / "Dividendy"           -> dividends + withholding tax
"""

import re
import openpyxl
import pandas as pd
import io
import csv
import requests
from datetime import datetime
from parsers.base import create_transaction


# --------------------------------------------------------------------------
# Localized sheet-name and column keywords (lowercased substrings)
# --------------------------------------------------------------------------

SHEET_KEYWORDS = {
    "closed":   ["closed positions", "zavřené pozice", "zavrene pozice"],
    "open":     ["open positions", "otevřené pozice", "otevrene pozice"],
    "activity": ["account activity", "aktivita na účtu", "aktivita na uctu"],
    "dividend": ["dividends", "dividendy"],
}

# Detection: any of these sheet substrings marks the workbook as eToro
ETORO_SHEET_INDICATORS = [
    "closed positions", "financial summary", "transactions report",
    "account activity", "account details",
    "zavřené pozice", "zavrene pozice", "aktivita na účtu", "aktivita na uctu",
    "finanční shrnutí", "financni shrnuti", "přehled účtu", "prehled uctu",
    "dividendy",
]

# Column resolvers: canonical key -> list of possible header substrings (lower)
COLS = {
    "action":      ["action", "akce"],
    "amount":      ["částka", "castka", "amount"],
    "units":       ["jednotky", "units"],
    "open_date":   ["datum otevření", "datum otevreni", "open date"],
    "close_date":  ["datum uzavření", "datum uzavreni", "close date"],
    "open_rate":   ["otevírací kurz", "oteviraci kurz", "open rate"],
    "close_rate":  ["uzavírací kurz", "uzaviraci kurz", "close rate"],
    "profit_eur":  ["zisk (eur)", "profit (eur)", "zisk eur"],
    "profit_usd":  ["zisk (usd)", "profit (usd)", "zisk usd"],
    "position_id": ["id pozice", "position id"],
    "isin":        ["isin"],
    "asset_type":  ["typ aktiva", "napište", "napiste", "type"],
    "date":        ["datum", "date"],
    "row_type":    ["napište", "napiste", "type"],
    "details":     ["podrobnosti", "details"],
    "current_rate": ["aktuální kurz", "aktualni kurz", "current rate", "market rate"],
    # Dividends sheet
    "div_date":    ["datum platby", "payment date"],
    "div_instrument": ["název nástroje", "nazev nastroje", "instrument name", "instrument"],
    "div_net_eur": ["čistá přijatá dividenda v (eur)", "net dividend received (eur)"],
    "div_net":     ["net dividends", "čistá přijatá dividenda", "cista prijata dividenda"],
    "div_currency": ["currency", "měna", "mena"],
    "div_wht":     ["částka srážkové daně", "castka srazkove dane", "withholding tax amount"],
}

# Activity row-type keywords (lowercased substrings of the "type" cell)
ACT_DEPOSIT    = ["deposit", "vklad"]
ACT_WITHDRAWAL = ["withdraw", "výběr", "vyber"]
ACT_DIVIDEND   = ["dividend", "dividenda"]
# Rows handled elsewhere (positions sheets) or not modeled — skipped in activity
# so they aren't misread as deposits/withdrawals. Checked before the rules above,
# which matters because e.g. "Poplatek za směnu měny při vkladu" contains "vklad".
ACT_SKIP = ["poplatek", "fee", "provize", "commission", "úprava", "uprava",
            "adjustment", "otevřená pozice", "otevrena pozice", "open position",
            "zisk/ztráta", "zisk/ztrata", "profit/loss", "rollover", "overnight"]


# --------------------------------------------------------------------------
# Small helpers
# --------------------------------------------------------------------------

def _safe_float(val, default=0.0) -> float:
    if val is None or val == "" or val == "-" or val == "N/A":
        return default
    try:
        s = str(val).strip()
        if s.startswith("(") and s.endswith(")"):
            return -float(s[1:-1].replace(",", ""))
        return float(s.replace(",", "").replace("$", "").replace("€", "").replace("%", ""))
    except (ValueError, TypeError):
        return default


def _parse_etoro_date(val) -> datetime:
    if isinstance(val, datetime):
        return val
    if isinstance(val, pd.Timestamp):
        return val.to_pydatetime()
    s = str(val).strip()
    for fmt in ["%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y",
                "%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y %H:%M:%S", "%m/%d/%Y"]:
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    try:
        return pd.to_datetime(s, dayfirst=True).to_pydatetime()
    except Exception:
        return datetime.now()


def _extract_ticker(instrument: str) -> str:
    """Pull a ticker from an eToro instrument label.

    eToro labels look like 'Taiwan Semiconductor ... - ADR (TSM)' or 'BTC/USD'
    or just 'Apple'. Prefer an explicit ticker in parentheses, then a SYM/CCY
    pair, then a short alnum token, else the first word.
    """
    if not instrument:
        return "UNKNOWN"
    s = str(instrument).strip()

    # 1) Ticker in parentheses, e.g. "(TSM)" or "(BRK.B)"
    m = re.search(r"\(([A-Za-z0-9.\-]{1,8})\)\s*$", s)
    if m:
        return m.group(1).upper()

    # 2) "SYM/CCY" pair, e.g. "TSM/USD" or "BTC/USD"
    if "/" in s:
        left = s.split("/")[0].strip()
        if 1 <= len(left) <= 8 and left.replace(".", "").replace("-", "").isalnum():
            return left.upper()

    # 3) Already a short ticker-ish token
    if len(s) <= 6 and s.replace(".", "").replace("-", "").isalnum():
        return s.upper()

    # 4) First word fallback
    return s.split()[0].upper() if s.split() else "UNKNOWN"


# --------------------------------------------------------------------------
# Sheet / header / column resolution
# --------------------------------------------------------------------------

def _find_sheet(wb, key):
    wanted = SHEET_KEYWORDS[key]
    for sn in wb.sheetnames:
        low = sn.strip().lower()
        if any(w in low for w in wanted):
            return wb[sn]
    return None


def _find_header_row(sheet, required_keys, max_scan=15):
    """Find the 1-based header row by requiring that every key in required_keys
    matches at least one cell. Returns (row_idx, header_list) or (-1, None)."""
    rows = list(sheet.iter_rows(values_only=True))
    for idx, row in enumerate(rows[:max_scan], start=1):
        if not row:
            continue
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        ok = True
        for key in required_keys:
            variants = COLS[key]
            if not any(any(v in cell for v in variants) for cell in cells):
                ok = False
                break
        if ok:
            header = [str(c).strip() if c is not None else "" for c in row]
            return idx, header
    return -1, None


def _rows_as_dicts(sheet, header_row_idx, header):
    rows = list(sheet.iter_rows(values_only=True))
    out = []
    for row in rows[header_row_idx:]:
        if not any(c not in (None, "") for c in row):
            continue
        d = {}
        for i, name in enumerate(header):
            if i < len(row):
                d[name] = row[i]
        out.append(d)
    return out


def _get(row_dict, key, default=None):
    """Get a value from a row dict by canonical column key (fuzzy header match)."""
    variants = COLS[key]
    for header, val in row_dict.items():
        hl = str(header).strip().lower()
        if any(v in hl for v in variants):
            return val
    return default


# --------------------------------------------------------------------------
# FX rate (USD -> EUR), derived from the statement when possible
# --------------------------------------------------------------------------

_live_rate_cache = {}

def _live_usd_eur() -> float:
    if "rate" in _live_rate_cache:
        return _live_rate_cache["rate"]
    try:
        resp = requests.get("https://api.frankfurter.app/latest?from=USD&to=EUR", timeout=5)
        if resp.status_code == 200:
            rate = resp.json().get("rates", {}).get("EUR", 0.92)
            _live_rate_cache["rate"] = rate
            return rate
    except Exception:
        pass
    _live_rate_cache["rate"] = 0.92
    return 0.92


def _derive_statement_rate(activity_rows) -> float:
    """Derive USD->EUR from a deposit whose details name an EUR amount.
    e.g. details '180.00 EUR CreditCard' with USD amount 208.13 -> 0.8648."""
    rates = []
    for r in activity_rows:
        rtype = str(_get(r, "row_type", "")).lower()
        if any(k in rtype for k in ACT_SKIP):
            continue
        if not any(k in rtype for k in ACT_DEPOSIT):
            continue
        details = str(_get(r, "details", "") or "")
        m = re.search(r"([\d.,]+)\s*EUR", details, re.IGNORECASE)
        usd = _safe_float(_get(r, "amount", 0))
        if m and usd > 0:
            eur = _safe_float(m.group(1))
            if eur > 0:
                rates.append(eur / usd)
    if rates:
        rates.sort()
        return rates[len(rates) // 2]  # median
    return _live_usd_eur()


# --------------------------------------------------------------------------
# Sheet parsers
# --------------------------------------------------------------------------

def _parse_closed(sheet, rate) -> list:
    txs = []
    idx, header = _find_header_row(sheet, ["action", "amount"])
    if idx == -1:
        idx, header = _find_header_row(sheet, ["position_id", "profit_eur"])
    if idx == -1:
        return txs
    for r in _rows_as_dicts(sheet, idx, header):
        instrument = str(_get(r, "action", "") or "")
        if not instrument or instrument.lower() in ("none", "nan", ""):
            continue
        open_date_raw = _get(r, "open_date")
        if not open_date_raw:
            continue
        ticker = _extract_ticker(instrument)
        units = _safe_float(_get(r, "units"))
        amount_usd = _safe_float(_get(r, "amount"))
        open_rate_usd = _safe_float(_get(r, "open_rate"))
        close_rate_usd = _safe_float(_get(r, "close_rate"))
        profit_eur = _get(r, "profit_eur")
        profit_eur = _safe_float(profit_eur) if profit_eur not in (None, "", "-") else _safe_float(_get(r, "profit_usd")) * rate
        pos_id = str(_get(r, "position_id", "") or "")

        if units == 0 and open_rate_usd > 0:
            units = amount_usd / open_rate_usd

        open_date = _parse_etoro_date(open_date_raw)
        close_raw = _get(r, "close_date")
        close_date = _parse_etoro_date(close_raw) if close_raw else open_date

        # Prices to EUR (rates are USD on the instrument)
        open_price_eur = open_rate_usd * rate
        close_price_eur = close_rate_usd * rate
        isin = _get(r, "isin")

        txs.append(create_transaction(
            date=open_date, ticker=ticker, tx_type="BUY", qty=units,
            price=open_price_eur, fee=0.0, position_id=pos_id, source="eToro", isin=isin))
        txs.append(create_transaction(
            date=close_date, ticker=ticker, tx_type="SELL", qty=units,
            price=close_price_eur, fee=0.0, realized_pnl=profit_eur,
            position_id=pos_id, source="eToro", isin=isin))
    return txs


def _parse_open(sheet, rate) -> list:
    txs = []
    idx, header = _find_header_row(sheet, ["action", "units"])
    if idx == -1:
        idx, header = _find_header_row(sheet, ["position_id", "open_rate"])
    if idx == -1:
        return txs
    for r in _rows_as_dicts(sheet, idx, header):
        instrument = str(_get(r, "action", "") or "")
        if not instrument or instrument.lower() in ("none", "nan", ""):
            continue
        open_date_raw = _get(r, "open_date")
        if not open_date_raw:
            continue
        ticker = _extract_ticker(instrument)
        units = _safe_float(_get(r, "units"))
        amount_usd = _safe_float(_get(r, "amount"))
        open_rate_usd = _safe_float(_get(r, "open_rate"))
        cur_rate_usd = _safe_float(_get(r, "current_rate"))
        pos_id = str(_get(r, "position_id", "") or "")
        if units == 0 and open_rate_usd > 0:
            units = amount_usd / open_rate_usd
        market_eur = (cur_rate_usd * rate) if cur_rate_usd > 0 else None
        txs.append(create_transaction(
            date=_parse_etoro_date(open_date_raw), ticker=ticker, tx_type="BUY",
            qty=units, price=open_rate_usd * rate, fee=0.0,
            market_price_at_export=market_eur, position_id=pos_id, source="eToro"))
    return txs


def _parse_activity(sheet, rate, activity_rows) -> list:
    """Deposits, withdrawals and dividends from the Account Activity ledger.
    Trade rows are intentionally ignored here (handled by the positions sheets)."""
    txs = []
    for r in activity_rows:
        date_raw = _get(r, "date")
        if not date_raw:
            continue
        dt = _parse_etoro_date(date_raw)
        rtype = str(_get(r, "row_type", "") or "").lower()
        amount_usd = _safe_float(_get(r, "amount"))
        details = str(_get(r, "details", "") or "")

        if any(k in rtype for k in ACT_SKIP):
            continue
        if any(k in rtype for k in ACT_DEPOSIT):
            m = re.search(r"([\d.,]+)\s*EUR", details, re.IGNORECASE)
            eur = _safe_float(m.group(1)) if m else abs(amount_usd) * rate
            txs.append(create_transaction(dt, "CASH", "DEPOSIT", eur, 1.0, source="eToro"))
        elif any(k in rtype for k in ACT_WITHDRAWAL):
            m = re.search(r"([\d.,]+)\s*EUR", details, re.IGNORECASE)
            eur = _safe_float(m.group(1)) if m else abs(amount_usd) * rate
            txs.append(create_transaction(dt, "CASH", "WITHDRAWAL", eur, 1.0, source="eToro"))
        elif any(k in rtype for k in ACT_DIVIDEND):
            ticker = _extract_ticker(str(_get(r, "details", "") or "CASH"))
            txs.append(create_transaction(dt, ticker, "DIVIDEND", abs(amount_usd) * rate, 1.0, source="eToro"))
    return txs


def _parse_dividends(sheet, rate) -> list:
    txs = []
    idx, header = _find_header_row(sheet, ["div_date", "div_instrument"])
    if idx == -1:
        return txs
    for r in _rows_as_dicts(sheet, idx, header):
        date_raw = _get(r, "div_date")
        if not date_raw:
            continue
        dt = _parse_etoro_date(date_raw)
        ticker = _extract_ticker(str(_get(r, "div_instrument", "") or ""))
        currency = str(_get(r, "div_currency", "EUR") or "EUR").upper()
        net_eur = _get(r, "div_net_eur")
        if net_eur not in (None, "", "-"):
            amount = _safe_float(net_eur)
        else:
            amount = _safe_float(_get(r, "div_net"))
            if currency != "EUR":
                amount *= rate
        if amount > 0:
            txs.append(create_transaction(dt, ticker, "DIVIDEND", amount, 1.0, source="eToro"))
        wht = _safe_float(_get(r, "div_wht"))
        if wht > 0:
            wht_eur = wht if currency == "EUR" else wht * rate
            txs.append(create_transaction(dt, ticker, "WITHHOLDING_TAX", wht_eur, 1.0, source="eToro"))
    return txs


# --------------------------------------------------------------------------
# CSV parser (Portfolio -> History export)
# --------------------------------------------------------------------------

def _parse_etoro_csv(filepath: str) -> list:
    rate = _live_usd_eur()
    with open(filepath, "r", encoding="utf-8") as f:
        lines = f.readlines()
    start = 0
    for i, line in enumerate(lines[:10]):
        if any(h in line for h in ["Position ID", "Action", "Open Date", "ID pozice"]):
            start = i
            break
    df = pd.read_csv(io.StringIO("".join(lines[start:])), delimiter=",", quoting=csv.QUOTE_MINIMAL)
    df.columns = [str(c).strip() for c in df.columns]
    txs = []

    def col(row, *names):
        for n in names:
            for c in row.index:
                if n.lower() in str(c).lower():
                    return row[c]
        return None

    for _, row in df.iterrows():
        instrument = str(col(row, "Action", "Akce", "Instrument") or "")
        action = instrument.lower()
        if not instrument:
            continue
        ticker = _extract_ticker(instrument)
        units = _safe_float(col(row, "Units", "Jednotky"))
        amount = _safe_float(col(row, "Amount", "Částka"))
        open_rate = _safe_float(col(row, "Open Rate", "Otevírací"))
        close_rate = _safe_float(col(row, "Close Rate", "Uzavírací"))
        profit_eur = col(row, "Profit (EUR)", "Zisk (EUR)")
        profit_eur = _safe_float(profit_eur) if profit_eur is not None else _safe_float(col(row, "Profit", "Zisk")) * rate
        pos_id = str(col(row, "Position ID", "ID pozice") or "")
        open_date = _parse_etoro_date(col(row, "Open Date", "Datum otevření", "Date"))
        close_raw = col(row, "Close Date", "Datum uzavření")
        close_date = _parse_etoro_date(close_raw) if pd.notna(close_raw) else open_date
        if units == 0 and open_rate > 0:
            units = amount / open_rate
        txs.append(create_transaction(open_date, ticker, "BUY", units, open_rate * rate,
                                      fee=0.0, position_id=pos_id, source="eToro"))
        if close_rate > 0:
            txs.append(create_transaction(close_date, ticker, "SELL", units, close_rate * rate,
                                          fee=0.0, realized_pnl=profit_eur, position_id=pos_id, source="eToro"))
    return txs


# --------------------------------------------------------------------------
# Public entry points
# --------------------------------------------------------------------------

def _parse_etoro_xlsx(filepath: str) -> list:
    wb = openpyxl.load_workbook(filepath, data_only=True)
    txs = []

    activity_sheet = _find_sheet(wb, "activity")
    activity_rows = []
    if activity_sheet:
        idx, header = _find_header_row(activity_sheet, ["date", "amount"])
        if idx != -1:
            activity_rows = _rows_as_dicts(activity_sheet, idx, header)

    rate = _derive_statement_rate(activity_rows)

    closed = _find_sheet(wb, "closed")
    if closed:
        txs.extend(_parse_closed(closed, rate))

    open_sheet = _find_sheet(wb, "open")
    if open_sheet:
        txs.extend(_parse_open(open_sheet, rate))

    if activity_rows:
        txs.extend(_parse_activity(activity_sheet, rate, activity_rows))

    dividends = _find_sheet(wb, "dividend")
    if dividends:
        txs.extend(_parse_dividends(dividends, rate))

    wb.close()
    return txs


def parse_etoro(filepath: str) -> list:
    """Parse an eToro export (XLSX Account Statement or CSV history)."""
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        return _parse_etoro_xlsx(filepath)
    elif ext == "csv":
        return _parse_etoro_csv(filepath)
    print(f"eToro parser: unsupported extension .{ext}")
    return []


def is_etoro_file(filepath: str) -> bool:
    """Detect an eToro export by characteristic (localized) sheet/column names."""
    ext = filepath.rsplit(".", 1)[-1].lower()
    if ext in ("xlsx", "xls"):
        try:
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            names = [sn.lower() for sn in wb.sheetnames]
            wb.close()
            return any(ind in sn for sn in names for ind in ETORO_SHEET_INDICATORS)
        except Exception:
            return False
    elif ext == "csv":
        try:
            with open(filepath, "r", encoding="utf-8") as f:
                sample = f.read(2048)
            return (("Position ID" in sample and "Open Rate" in sample)
                    or ("ID pozice" in sample and "Otevírací" in sample)
                    or "eToro" in sample)
        except Exception:
            return False
    return False
